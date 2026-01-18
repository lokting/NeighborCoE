#合并count文件，并计算RPKM
import os
import openpyxl as xl
import pandas as pd
from bioinfokit.analys import norm
import math
from itertools import permutations
from scipy.stats import pearsonr

#定义函数 对输入进行 加1后log2 的处理
def log_RPKM(x):
    out = math.log2(x+1)
    return out

input_count_file_path = r"input/Count_file"
input_excel_path = r"tmp_output/gff.xlsx" 
chr_list_file = r'input/chr_list.txt'
pla_list_file = r'input/pla_list.txt'

output_gene_sample_raw = r"tmp_output/gene_sample_raw.xlsx"
output_RPKM = r"tmp_output/RPKM.xlsx"
output_RPKM_log2 = r"tmp_output/RPKM_log2.xlsx"
output_gene_PCC_txt = r"tmp_output/gene_PCC.txt"

def read_list_from_file(file_path):
    with open(file_path, 'r') as file:
        return [line.strip() for line in file]
    
#染色体、质粒的sheetID
chr_list = read_list_from_file(chr_list_file)
pla_list = read_list_from_file(pla_list_file)

# bowtie2<70% 的样本（需要删除） 
# localtag_dele = ["SRR2444787","SRR2444788","SRR2444790","SRR2444791","SRR2444793","SRR2444794"] 
localtag_dele_file = r'input/local_tag_dele.txt'
localtag_dele = [line.strip() for line in read_list_from_file(localtag_dele_file)]


#构造{sample:{localtag:count}}字典表
dict_sample = {}
for root, dirs, files in os.walk(input_count_file_path):#更改对应路径  文件夹内文件批量循环   
   for name in files:

    dict_local_tag = {}

    sample_name = name.split("_")[0]
    txt_file_path = os.path.join(root, name)

    f = open(txt_file_path)
    for line in f.readlines():
        if line.startswith("__"):
            continue
        else:
            line= line.split("\t")
            local_tag = line[0]
            read_count = line[1]
            read_count = read_count.replace('\n','')
            dict_local_tag[local_tag] = int(read_count)
    dict_sample[sample_name] = dict_local_tag


# 读取Excel数据 构造{length{localtag:length_num}} 的字典表
excel_local_tag = {}
excel_dict = {}
workbook = xl.load_workbook(input_excel_path)
# sheeet_names_1 = ["NC_016810.1_+","NC_016810.1_+"]
for sheet_name in workbook.sheetnames:
# for sheet_name in sheeet_names_1:
    worksheet = workbook[sheet_name]
    num = 1
    finial = worksheet.max_row #表格数量
    while num<finial:
        num = num + 1

        # 读取值
        cell_Local_tag = worksheet["E"+str(num)]
        excel_Local_tag = cell_Local_tag.value
        cell_gene_length = worksheet["F"+str(num)]
        excel_gene_length = cell_gene_length.value

        excel_local_tag[excel_Local_tag] = int(excel_gene_length)
excel_dict["length"] = excel_local_tag


#合并计数文件结果并保存
dict_sample.update(excel_dict) #将样本的dict和长度的dict合并成一个字典表
frame = pd.DataFrame(dict_sample)#构造dataframe 外键为列名 内建为行名
frame.index.name = "Local_tag"  #将行名设为索引


#删除质粒部分
local_tag_del_list = []
# for sheet_name in workbook.sheetnames:
for sheet_name in pla_list:
    worksheet = workbook[sheet_name]
    num = 1
    finial = worksheet.max_row #表格数量
    while num<finial:
        num = num + 1
        # 读取值
        cell_Local_tag = worksheet["E"+str(num)]
        excel_Local_tag = cell_Local_tag.value
        local_tag_del_list.append(excel_Local_tag)
frame = frame.drop(local_tag_del_list)

#保存原始结果
frame.to_excel(output_gene_sample_raw)

#计算RPKM并保存结果
nm = norm()
nm.rpkm(df=frame, gl='length') #指定基因长度列
rpkm_df = nm.rpkm_norm  #计算RPKM
rpkm_df.to_excel(output_RPKM)

#计算log2(RPKM+1)并保存结果
count_RPKM_log2 = rpkm_df.applymap(log_RPKM) #引入log2(RPKM+1)函数 对所有值处理
count_RPKM_log2.to_excel(output_RPKM_log2)
print("count_RPKM_log2")
print(count_RPKM_log2.shape)


#以上为所有样本数据
#接下来开始剔除质量比较差的数据  （在此以 bowtie2<70% 者剔除 ）
# #以列表形式存储需要剔除的样本名
count_RPKM_log2_new = count_RPKM_log2.drop(labels = localtag_dele ,axis = 1,inplace = False)
print("count_RPKM_log2_new")
print(count_RPKM_log2_new.shape)


#计算pearson系数
def get_pearsonr_num(dataframe,excel_path,chr_list):  #定义计算pearson系数函数
    pearsonr_list = []
    #各自读取四张表local_tag并组合
    workbook = xl.load_workbook(excel_path)
    # for sheet_name in workbook.sheetnames:
    for sheet_name in chr_list:
        excel_Local_tag_list = []
        worksheet = workbook[sheet_name]
        num = 1
        finial = worksheet.max_row #表格数量
        while num<finial:
            num = num + 1
            # 读取值
            cell_Local_tag = worksheet["E"+str(num)]
            excel_Local_tag_list.append(cell_Local_tag.value)

        #frame_index = dataframe.index.tolist()#df.index 返回的是 index 对象列表，需转换为普通列表格式时用 tolist() 方法
        for index_name in permutations(excel_Local_tag_list, 2): #来自 itertools 模块的函数 permutations(input, x) 返回的是一个长度为 r 的所有可能排列(包括a-b b-a) 用for循环依次读取
                local_tag_1 = index_name[0]
                local_tag_2 = index_name[1]
                data_1 = dataframe.loc[local_tag_1]
                data_2 = dataframe.loc[local_tag_2]
                pearsonr_num = pearsonr(data_1,data_2) #计算两个数据之间的pearson系数，输出:(r, p) r:相关系数[-1，1]之间 p:相关系数显著性
                pearsonr_num_value = pearsonr_num[0]
                pearsonr_p_value = pearsonr_num[1]
                pearsonr_list.append(str(local_tag_1) + "-" + str(local_tag_2) + "\t" + str(pearsonr_num_value)+ "\t" + str(pearsonr_p_value))
    return pearsonr_list

#调用函数 计算染色体与质粒 所有localtag组合的PCC和p值  （函数设计了列表 可以叠加 所以不用返回 后期使用列表就行）
pearsonr_list = get_pearsonr_num(count_RPKM_log2_new,input_excel_path,chr_list)

#将PCC_list存入txt
with open(output_gene_PCC_txt,"a+",encoding="utf-8") as pearson_num_file:
    for data in pearsonr_list:
        pearson_num_file.write(data + "\n") # \n 换行

